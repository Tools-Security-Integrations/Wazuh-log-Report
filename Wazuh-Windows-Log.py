#Created by Vaibhav Handekar
import json
import pandas as pd
import chardet
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import paramiko
import os
from datetime import datetime

# Load SSH credentials from environment variables or configuration file
LINUX_HOST = os.getenv('LINUX_HOST', 'IP Address (Replace with IP)')  # Replace with actual value or keep environment variable
LINUX_USER = os.getenv('LINUX_USER', 'wazuh-user')  # Replace with actual value or keep environment variable
LINUX_PASSWORD = os.getenv('LINUX_PASSWORD', 'wazuh')  # Replace with actual value or keep environment variable

# Define paths
current_working_dir = os.getcwd()
EXCEL_FILE_PATH = os.path.join(current_working_dir, 'ossec-alerts.xlsx')

# Define the list of columns and fields to include in the DataFrame
columns = ['Timestamp', 'Agent IP', 'Agent Name', 'IP', 'Port', 'Full log', 'Rule Level', 'Description', 'Event','Image', 'Query Name', 'Mitre Id', 'Mitre Tactic', 'Mitre Technique',  'Severity Value',
           'EventID']
fields = ['timestamp', 'agent.ip', 'agent.name', 'data.win.eventdata.ipAddress', 'data.win.eventdata.ipPort', 'full_log', 'rule.level', 'rule.description', 'syscheck.event', 'data.win.eventdata.image', 'data.win.eventdata.queryName', 'rule.mitre.id', 'rule.mitre.tactic', 'rule.mitre.technique', 'data.win.system.severityValue',
          'data.win.system.eventID']

def transfer_latest_json():
    ssh_client = paramiko.SSHClient()
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh_client.connect(LINUX_HOST, username=LINUX_USER, password=LINUX_PASSWORD)

    sftp = ssh_client.open_sftp()

    try:
        # Get current year, month, and date
        current_year = datetime.now().strftime('%Y')
        current_month = datetime.now().strftime('%b')
        current_date = datetime.now().strftime('%d')

        # Remote and local file paths
        remote_file_path = f"/var/ossec/logs/alerts/{current_year}/{current_month}/ossec-alerts-{current_date}.json"
        local_file_path = f'ossec-alerts-{current_date}.json'

        # Debugging output
        print(f"Attempting to transfer file from: {remote_file_path} to {local_file_path}")

        # Transfer the file
        sftp.get(remote_file_path, local_file_path)
        return local_file_path
    except Exception as e:
        print(f"Error transferring file: {e}")
        return None
    finally:
        sftp.close()
        ssh_client.close()

def process_json_to_excel(local_json_path):
    if local_json_path:
        # Detect the encoding of the JSON file
        with open(local_json_path, 'rb') as f:
            result = chardet.detect(f.read())

        # Read and parse the JSON file with the detected encoding
        dfs = []
        with open(local_json_path, 'r', encoding=result['encoding']) as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        json_obj = json.loads(line)
                        data = {
                            columns[0]: json_obj.get(fields[0]).split('.')[0].replace('T',' '),
                            columns[1]: json_obj.get('agent', {}).get('ip'),
                            columns[2]: json_obj.get('agent', {}).get('name'),
                            columns[3]: json_obj.get('data', {}).get('win', {}).get('eventdata', {}).get('ipAddress'),
                            columns[4]: json_obj.get('data', {}).get('win', {}).get('eventdata', {}).get('ipPort'),
                            columns[5]: json_obj.get(fields[5]),
                            columns[6]: json_obj.get('rule', {}).get('level'),
                            columns[7]: json_obj.get('rule', {}).get('description'),
                            columns[8]: json_obj.get('syscheck', {}).get('event'),
                            columns[9]: json_obj.get('data', {}).get('win', {}).get('eventdata', {}).get('image'),
                            columns[10]: json_obj.get('data', {}).get('win', {}).get('eventdata', {}).get('queryName'),
                            columns[11]: ', '.join(json_obj.get('rule', {}).get('mitre', {}).get('id', [])),
                            columns[12]: ', '.join(json_obj.get('rule', {}).get('mitre', {}).get('tactic', [])),
                            columns[13]: ', '.join(json_obj.get('rule', {}).get('mitre', {}).get('technique', [])),
                            columns[14]: json_obj.get('data', {}).get('win', {}).get('system', {}).get('severityValue'),
                            columns[15]: json_obj.get('data', {}).get('win', {}).get('system', {}).get('eventID')
                            
                            
                        }
                        df = pd.DataFrame([data])
                        dfs.append(df)
                    except json.JSONDecodeError:
                        print(f"Error decoding JSON line: {line}")

        # Concatenate all the DataFrames into a single DataFrame
        if dfs:
            df = pd.concat(dfs, ignore_index=True)

            # Write the DataFrame to an Excel file
            if os.path.exists(EXCEL_FILE_PATH):
                with pd.ExcelWriter(EXCEL_FILE_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    startrow = writer.sheets['Sheet1'].max_row
                    df.to_excel(writer, index=False, header=False, startrow=startrow)
            else:
                df.to_excel(EXCEL_FILE_PATH, index=False)

            # Load the workbook and access the active sheet
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet = workbook.active

            # Set the width for each column
            column_widths = {
                'A': 19.86, 'B': 14.29, 'C': 17.29, 'D': 14.29, 'E':8.15, 'F': 39.29, 'G': 11.29,
                'H': 39.29, 'I': 8.57, 'J': 39.29, 'K': 39.29, 'L': 15.29, 'M': 25.29, 'N': 35.29,
                'O': 14.43, 'P': 10.29
            }
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Set text wrapping for specific columns
            wrap_columns = ['G']  # Columns to apply text wrapping (Image, Query Name)
            for col in wrap_columns:
                for cell in sheet[col]:
                    cell.alignment = Alignment(wrap_text=True)

            # Freeze the top row and the first two columns
            sheet.freeze_panes = 'D2'

            # Save the workbook
            workbook.save(EXCEL_FILE_PATH)
        else:
            print("No valid JSON data found.")

def main():
    local_json_path = transfer_latest_json()
    process_json_to_excel(local_json_path)

if __name__ == "__main__":
    main()
